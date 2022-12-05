#!python 3

"""
Python leetcode tracker by: Auden Woolfson

TODO: handle abs/rel paths with os.path.relpath (isabs),
use dictionary to optimize arg parsing,
debug waiting for loading screen to finish (inconsistent),
write README,
test xl more extensively, add formatting stuff?
"""

import sys
import pyinputplus as pyip # gets user inputs
import requests # used for validating URLs
from bs4 import BeautifulSoup # parses html
import openpyxl as pyxl # reads and writes to xlsx (Microsoft Excel) files
import re # regular expressions used for filepath and URL
import os # used to validate file paths

from selenium import webdriver # selenium webdriver used for downloading HTML
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
def main():
    
    #user input variables
    
    leetcodeURLInput = ""
    xlFilepathInput = ""
    solutionInput = ""
    notesInput = ""
    xlSheetName = "Sheet1" # defaul sheet name is Sheet1, --sheet argument can specify
    
    #parse arguments
    
    args = sys.argv
    for i in range(len(args)-1): #-1 to prevent i+1 index from being out of range
        if args[i] == "-u":
            leetcodeURLInput = args[i + 1]
        if args[i] == "-x":
            xlFilepathInput = args[i + 1]
            relativeXlPath = False
        if args[i] == "-s":
            solutionInput = args[i + 1]
        if args[i] == "-n":
            notesInput = args[i + 1]
        if args[i] == "-r":
            xlWorkbookFilepath = args[i + 1]
            relativeXlPath = True
        if args[i] == "--sheet":
            xlSheetName = args[i + 1]
    
    solution = solutionInput
    notes = notesInput
            
    # validate leetcode URL with regex to avoid errors with request, will ask for new addres again later if request fails
    
    leetcodeURLRegex = re.compile(r'https:\/\/leetcode.com\/problems\/\S+')
    
    # example: https://leetcode.com/problems/two-sum/
    
    leetcodeURL = getLeetcodeURLInput(leetcodeURLRegex, leetcodeURLInput) 
    
    # validate xl file path by extension and path validity
    
    xlRegex = re.compile(r'.*\.xlsx?$')
    xlWorkbookFilepath = getXlFilepathInput(xlRegex, xlFilepathInput)
    
    # validate URL again using the web with standard python requests
    
    leetcodeRes = requests.get(leetcodeURL)
    print(f'pre web validation: {leetcodeRes}')
    
    leetcodeURL = tryRequest(leetcodeURLRegex, leetcodeRes, leetcodeURL)
    if leetcodeRes == False:
        print("""
...
""")
        exit()
    
    print(f'post web validation: {leetcodeRes}')
    #print(leetcodeRes.text)
    
    # get HTML document in html/leetcodeproblem.html (relative path, previous content cleared)
    
    options = webdriver.FirefoxOptions()
    options.add_argument('--headless') # headless browser does not open a window
    
    browser = webdriver.Firefox(options = options)
    browser.get(leetcodeURL)
    WebDriverWait(browser, 10).until(
        ec.presence_of_element_located((By.CLASS_NAME, 'css-isal7m')))
    WebDriverWait(browser, 10)
    
    # expand the related topics tab'
    
    topicsElementClass = browser.find_elements(By.CLASS_NAME, 'css-isal7m')
    expandTopicsElement = topicsElementClass[1] # topics element is second on page in class
    expandTopicsElement.click() # click it to expand
    
    # save HTML and close browser
    
    sourceHTML = browser.page_source
    browser.quit() 
   
    with open("html/leetcodeproblem.html",'r+') as file: # clear the file
        file.truncate(0)
        file.close()
    
    localHTMLFile = open("html/leetcodeproblem.html", "w")
    localHTMLFile.write(sourceHTML) # write the HTML
    localHTMLFile.close()
    
    # prettify HTML with BeautifulSoup from bs4
    
    localHTMLFile = open("html/leetcodeproblem.html", "r+")
    index = localHTMLFile.read()
    HTMLBeautifulSoup = BeautifulSoup(index, 'lxml')
    localHTMLFile.truncate(0)
    prettyHTML = HTMLBeautifulSoup.body.prettify()
    localHTMLFile.write(prettyHTML)
    localHTMLFile.close()
    
    # get the relevant elements
    
    # problem name: class="css-v3d350"
    # difficulty: class="css-14oi08n"
    # topics: class="css-vrmejz"
    
    # parse the relevant elements
    
    problemNameElement = HTMLBeautifulSoup.find(class_ = "css-v3d350")
    difficultyElement = HTMLBeautifulSoup.find(class_ = "css-14oi08n")
    relatedTopicsElement = HTMLBeautifulSoup.find(class_ = "css-vrmejz") # after expansion

    problemTitle = problemNameElement.text
    difficulty = difficultyElement.text
    
    numberRegex = re.compile(r'\d+')
    problemNumberMatch = re.match(numberRegex, problemTitle)
    problemNumber = problemNumberMatch.group()
    
    for i in range(len(problemTitle)):
        if problemTitle[i] == '.':
                problemName = problemTitle[i + 2:]
    
    relatedTopicsChildren = relatedTopicsElement.findChildren()
    relatedTopics = []
    num = 1
    
    print(f'{problemNumber} {problemName} {difficulty}')
    for index, topic in enumerate(relatedTopicsChildren):
        if index % 2 == 0:
            relatedTopics.append(topic.text)
            print(f'topic {num}. {topic.text}')
            num += 1
    
    # this needs to be tested
    
    xlWorkbook = pyxl.load_workbook(xlWorkbookFilepath)
    xlsheetnames = xlWorkbook.sheetnames
    if len(xlsheetnames) == 1:
        print(f'1 sheet found: {xlsheetnames[0]}')
        xlSheet = xlWorkbook[xlsheetnames[0]]
    else:
        try:
            xlSheet = xlWorkbook[xlSheetName]
        except:
            print("""
error: there was a problem finding the sheet Sheet1 or [--sheet] in the xl workboook                
""")
            quit()
    
    print(f'{xlSheet.max_row}') 
    for rowNum in range(2, xlSheet.max_row + 2):
        print(f'checking row {rowNum}')
        if xlSheet.cell(row = rowNum, column = 2).value == None or xlSheet.cell(row = rowNum, column = 2).value == problemName:
            xlSheet.cell(row = rowNum, column = 2).value = problemName
            xlSheet.cell(row = rowNum, column = 3).value = difficulty
            xlSheet.cell(row = rowNum, column = 4).value = (', '.join(relatedTopics))
            xlSheet.cell(row = rowNum, column = 5).value = solution
            xlSheet.cell(row = rowNum, column = 6).value = notes
            print("workbook updated")
            xlWorkbook.save(xlWorkbookFilepath)
            quit()
    
    
    
# HELPER METHODS

def getLeetcodeURLInput(methodRegex, URLInput):
    # validates leetcode URLs based on regex (no internet required)
    # example: https://leetcode.com/problems/two-sum/
    while not re.match(methodRegex, URLInput):
        URLInput = pyip.inputRegex(
            prompt = """
Please enter the URL of the leetcode problem you would like to enter
(any previous URL may have been entered incorrectly)
""",
            regex = methodRegex
        )
    return URLInput

def getXlFilepathInput(methodRegex, pathInput):
    # validates xl file path by extension and path validity
    while not os.path.isfile(pathInput) or not re.match(methodRegex, pathInput):
        pathInput = pyip.inputStr(
            prompt = """
Please enter the absolute file path to the .xlsx file you would like to make an entry to
(any previous path may have been entered incorreclty)
"""
        )
    return pathInput

def tryRequest(methodRegex, res, leetcodeURL):
    # recursive method for validating URL through requests
    try:
        res.raise_for_status()
        return leetcodeURL
    except Exception as exc:
        print(f'There was a problem with the URL: {exc}')
        tryAgain = pyip.inputYesNo(
            prompt = "would you like to try a different URL? (yes/no)"
        )
        if not tryAgain:
            return False # used for validating success of method
    leetcodeURL = getLeetcodeURLInput(methodRegex, "")
    return tryRequest(methodRegex, requests.get(leetcodeURL), leetcodeURL)

main()