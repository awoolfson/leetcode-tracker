#! python3

"""
This is the secondary script in the leetcode tracker project by Auden Woolfson    
Please see the README.md file for more information on the first script,
trackerget.py should be seen as a pre-requisite to this

The purpose of this script is to randomly select a problem for the user to solve from
the spreadsheet created by trackerget.py. The user can specify the topics and difficulty
"""

from trackerget import *

import sys
import pyinputplus as pyip # gets user inputs
import requests # used for validating URLs
from bs4 import BeautifulSoup # parses html
import openpyxl as pyxl # reads and writes to xlsx (Microsoft Excel) files
import re # regular expressions used for filepath and URL
import os # used to validate file paths

def main():
    if __name__ == "__main__":
        
        args = sys.argv
        relatedTopicsInput = []
        xlFilepathInput = ""
        relatedTopicsInput = "Dynamic Programming, Math" # just defaults for example
        problemDifficultyInput = "Easy, Medium, Hard"
        xlSheetName = "Sheet1" # defaul sheet name is Sheet1, --sheet argument can specify
        
        for i in range(len(args) - 1):
            if args[i] == "-r":
                relatedTopicsInput = args[i + 1]
            if args[i] == "-x":
                xlFilepathInput = args[i + 1]
            if args[i] == "-d":
                problemDifficultyInput = args[i + 1]
            if args == "--sheet":
                xlSheetName = args[i + 1]
                
        relatedTopicsList = relatedTopicsInput.split(",")
        relatedTopicsDict = {}
        for index, topic in enumerate(relatedTopicsList):
            relatedTopicsDict[topic] = True
            
        problemDifficultyList = problemDifficultyInput.split(",")
        problemDifficultyDict = {}
        for index, difficulty in enumerate(problemDifficultyList):
            problemDifficultyDict[difficulty] = True
        
        xlRegex = re.compile(r'.*\.xlsx?$')
        xlWorkbookFilepath = getXlFilepathInput(xlRegex, xlFilepathInput)
        
        xlWorkbook = pyxl.load_workbook(xlWorkbookFilepath)
        xlsheetnames = xlWorkbook.sheetnames
        if len(xlsheetnames) == 1:
            print(f'1 sheet found: {xlsheetnames[0]}')
            xlSheet = xlWorkbook[xlsheetnames[0]]
        else:
            try:
                xlSheet = xlWorkbook[xlSheetName]
            except:
                print("""
    error: there was a problem finding the sheet Sheet1 or [--sheet] in the xl workboook                
    """)
                quit()
                
        diffucultyMatches = 0
        topicsMatches = 0
                
        for rowNum in range(3, xlSheet.max_row + 2):
            # here I will keep track of how many matches there are for 
            # difficulty, related topics
main()